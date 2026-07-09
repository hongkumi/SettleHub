from flask import Flask, render_template, request, jsonify, redirect, url_for
import openpyxl
import os
import re
import sqlite3
from datetime import datetime
from pathlib import Path

app = Flask(__name__)

EXCEL_PATH = os.environ.get(
    'EXCEL_PATH',
    r'C:\Users\Hong\Downloads\★★제휴정산_입금내역★★.xlsx'
)
DB_PATH = Path(__file__).parent / 'settlehub.db'

# 두발히어로 단가
DUBBAL_UNIT_PRICE = 10  # 건당 10원 (부가세 별도)
VAT_RATE = 0.10

_cache = {}

# 파트너명 → 시트명 매핑
PARTNER_SHEETS = {
    'KG이니시스':       '1.KG이니시스',
    'GTF':             '3.GTF',
    '리터놀':           '4.리터놀',
    '버클 (선발행)':    '5. 버클',
    '비플':             '6.비플',
    '야마토':           '7.야마토(YSM)',
    '채널코퍼레이션':   '11.채널코퍼레이션',
    '체큐리티(나이스)': '12.체큐리티(나이스)',
    '카카오모빌리티':   '13.카카오모빌리티',
    '토스페이먼츠':     '14.토스페이먼츠',
    '패스트박스':       '15.패스트박스',
    '스토어X':          '17.스토어X',
    '두발히어로':       '19.두발히어로',
    '엑심베이':         '20.엑심베이',
}

# ── DB 초기화 ──────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS dubbal_records (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                ym          TEXT NOT NULL UNIQUE,  -- YYYY-MM
                count       INTEGER NOT NULL DEFAULT 0,
                invoice_date TEXT,                 -- 세금계산서 발행일
                payment_date TEXT,                 -- 입금일
                note        TEXT,
                paid        INTEGER NOT NULL DEFAULT 0,  -- 0/1
                created_at  TEXT DEFAULT (datetime('now','localtime')),
                updated_at  TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        # 엑셀 기존 데이터 초기 적재 (이미 있으면 skip)
        seed = [
            ('2026-04', 171812, '2026-05-31', None,   '24년10월~26년4월 정산서', 0),
            ('2026-05', 13690,  '2026-06-30', None,   None, 0),
            ('2026-06', 8650,   '2026-07-31', None,   None, 0),
        ]
        for row in seed:
            conn.execute("""
                INSERT OR IGNORE INTO dubbal_records (ym, count, invoice_date, payment_date, note, paid)
                VALUES (?, ?, ?, ?, ?, ?)
            """, row)
        conn.commit()


# ── 공통 파싱 유틸 ────────────────────────────────────────
def parse_month_label(raw):
    if not raw:
        return ''
    m = re.search(r'(\d{4})\.(\d+)월', str(raw))
    if m:
        return f"{m.group(2)}월"
    m2 = re.search(r'(\d+)월', str(raw))
    if m2:
        return f"{m2.group(1)}월"
    return str(raw)[:10]


def load_status_data():
    if 'status' in _cache:
        return _cache['status']

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['입금여부']

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {'months': [], 'partners': []}

    header = rows[0]
    months = [parse_month_label(header[i]) for i in range(1, 13)]

    partners = []
    for row in rows[1:]:
        if not row[0] or not str(row[0]).strip():
            continue
        name = str(row[0]).strip()
        note = str(row[13]).strip() if len(row) > 13 and row[13] else ''
        note = '' if note == 'None' else note

        statuses = []
        for i in range(1, 13):
            val = row[i] if i < len(row) else None
            v = str(val).strip() if val is not None else ''
            if not v or v == 'None':
                statuses.append({'type': 'pending', 'label': ''})
            elif '✔' in v:
                statuses.append({'type': 'paid', 'label': '✔'})
            elif v.lower() == 'x':
                statuses.append({'type': 'refused', 'label': '✕'})
            else:
                statuses.append({'type': 'na', 'label': v[:8]})

        partners.append({
            'name': name,
            'statuses': statuses,
            'note': note,
            'has_sheet': name in PARTNER_SHEETS,
        })

    result = {'months': months, 'partners': partners}
    _cache['status'] = result
    return result


def load_partner_detail(partner_name):
    sheet_name = PARTNER_SHEETS.get(partner_name)
    if not sheet_name:
        return None

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    records = []
    for row in rows:
        if not any(row):
            continue
        month_val = str(row[0]).strip() if row[0] else ''
        if not re.match(r'^\d+월$', month_val):
            continue

        amounts = []
        texts = []
        for cell in row[1:]:
            if cell is None:
                continue
            try:
                amt = float(cell)
                if amt > 0:
                    amounts.append(int(amt))
            except (TypeError, ValueError):
                s = str(cell).strip()
                if s and s != 'None':
                    texts.append(s)

        total = sum(amounts) if amounts else 0
        records.append({
            'month': month_val,
            'amounts': amounts,
            'texts': texts,
            'total': total,
        })

    return records


# ── 두발히어로 유틸 ───────────────────────────────────────
def calc_dubbal(count):
    supply = count * DUBBAL_UNIT_PRICE
    vat    = round(supply * VAT_RATE)
    total  = supply + vat
    return supply, vat, total


def get_dubbal_records():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM dubbal_records ORDER BY ym"
        ).fetchall()
    return [dict(r) for r in rows]


def ym_to_label(ym):
    """'2026-07' → '2026년 7월'"""
    try:
        y, m = ym.split('-')
        return f"{y}년 {int(m)}월"
    except Exception:
        return ym


# ── 라우트: 메인 ─────────────────────────────────────────
@app.route('/')
def index():
    data = load_status_data()
    partners = data['partners']
    months = data['months']

    now_month = datetime.now().month
    cur_idx = min(now_month - 1, 11)

    stats = []
    for i, m in enumerate(months):
        paid    = sum(1 for p in partners if p['statuses'][i]['type'] == 'paid')
        refused = sum(1 for p in partners if p['statuses'][i]['type'] == 'refused')
        pending = sum(1 for p in partners if p['statuses'][i]['type'] == 'pending')
        na      = sum(1 for p in partners if p['statuses'][i]['type'] == 'na')
        stats.append({'month': m, 'paid': paid, 'refused': refused,
                      'pending': pending, 'na': na})

    return render_template('index.html',
                           months=months,
                           partners=partners,
                           stats=stats,
                           cur_idx=cur_idx,
                           total=len(partners))


@app.route('/partner/<path:name>')
def partner_detail(name):
    if name == '두발히어로':
        return redirect(url_for('dubbal_index'))

    data = load_status_data()
    partner = next((p for p in data['partners'] if p['name'] == name), None)
    if not partner:
        return "파트너를 찾을 수 없습니다.", 404

    records = load_partner_detail(name) or []
    grand_total = sum(r['total'] for r in records)
    return render_template('partner.html',
                           partner=partner,
                           months=data['months'],
                           records=records,
                           grand_total=grand_total)


# ── 라우트: 두발히어로 ────────────────────────────────────
@app.route('/dubbal', methods=['GET', 'POST'])
def dubbal_index():
    msg = None
    if request.method == 'POST':
        ym           = request.form.get('ym', '').strip()
        count_raw    = request.form.get('count', '').strip()
        invoice_date = request.form.get('invoice_date', '').strip() or None
        payment_date = request.form.get('payment_date', '').strip() or None
        note         = request.form.get('note', '').strip() or None

        if not re.match(r'^\d{4}-\d{2}$', ym):
            msg = ('danger', 'YYYY-MM 형식으로 입력하세요.')
        elif not count_raw.isdigit():
            msg = ('danger', '건수는 숫자로 입력하세요.')
        else:
            count = int(count_raw)
            with get_db() as conn:
                conn.execute("""
                    INSERT INTO dubbal_records (ym, count, invoice_date, payment_date, note)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(ym) DO UPDATE SET
                        count        = excluded.count,
                        invoice_date = excluded.invoice_date,
                        payment_date = excluded.payment_date,
                        note         = excluded.note,
                        updated_at   = datetime('now','localtime')
                """, (ym, count, invoice_date, payment_date, note))
                conn.commit()
            msg = ('success', f'{ym_to_label(ym)} {count:,}건 저장 완료.')

    records = get_dubbal_records()
    enriched = []
    for r in records:
        supply, vat, total = calc_dubbal(r['count'])
        enriched.append({**r, 'supply': supply, 'vat': vat, 'total': total,
                         'label': ym_to_label(r['ym'])})

    return render_template('dubbal.html', records=enriched, msg=msg,
                           unit_price=DUBBAL_UNIT_PRICE)


@app.route('/dubbal/<int:rid>/delete', methods=['POST'])
def dubbal_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM dubbal_records WHERE id = ?", (rid,))
        conn.commit()
    return redirect(url_for('dubbal_index'))


@app.route('/dubbal/<int:rid>/toggle-paid', methods=['POST'])
def dubbal_toggle_paid(rid):
    with get_db() as conn:
        conn.execute(
            "UPDATE dubbal_records SET paid = 1 - paid WHERE id = ?", (rid,))
        conn.commit()
    return redirect(url_for('dubbal_index'))


@app.route('/dubbal/invoice/<ym>')
def dubbal_invoice(ym):
    """인쇄용 청구서"""
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM dubbal_records WHERE ym = ?", (ym,)
        ).fetchone()
    if not row:
        return "해당 월 데이터가 없습니다.", 404

    r = dict(row)
    supply, vat, total = calc_dubbal(r['count'])
    label = ym_to_label(ym)
    issue_date = r['invoice_date'] or datetime.now().strftime('%Y-%m-%d')

    return render_template('dubbal_invoice.html',
                           r=r, supply=supply, vat=vat, total=total,
                           label=label, issue_date=issue_date,
                           unit_price=DUBBAL_UNIT_PRICE)


# ── 기타 ─────────────────────────────────────────────────
@app.route('/api/reload')
def reload_cache():
    _cache.clear()
    return jsonify({'ok': True, 'message': '캐시가 초기화되었습니다.'})


if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5001)
